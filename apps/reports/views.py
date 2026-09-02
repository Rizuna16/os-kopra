import csv
from datetime import datetime, time
from decimal import Decimal
from io import BytesIO

from django.db.models import Count, F, Sum, Q, Value
from django.db.models.functions import Coalesce
from django.http import HttpResponse
from django.utils import timezone

from django.shortcuts import get_object_or_404
from rest_framework import status as drf_status
from rest_framework.exceptions import ValidationError
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from apps.business.models import Location
from apps.customer.models import Customer
from apps.employee.models import Employee
from apps.finance.models import Expense, Journal, JournalEntry
from apps.product.models import Product, Variant
from apps.purchasing.models import PurchaseOrder, PurchaseOrderLine
from apps.sales.models import Sale, SaleLine, CashierShift
from apps.supplier.models import Supplier
from apps.promotion_loyalty.models import Promotion
from apps.authentication.permissions import BusinessAccessMixin

try:
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


def parse_date_params(request):
    raw_from = request.query_params.get("date_from")
    raw_to = request.query_params.get("date_to")
    dt_from = None
    dt_to = None
    tz = timezone.get_current_timezone()
    if raw_from:
        try:
            d = datetime.strptime(raw_from, "%Y-%m-%d").date()
        except ValueError:
            raise ValidationError("Invalid date_from. Expected YYYY-MM-DD.")
        dt_from = datetime.combine(d, time.min, tzinfo=tz)
    if raw_to:
        try:
            d = datetime.strptime(raw_to, "%Y-%m-%d").date()
        except ValueError:
            raise ValidationError("Invalid date_to. Expected YYYY-MM-DD.")
        dt_to = datetime.combine(d, time(23, 59, 59, 999999), tzinfo=tz)
    if dt_from and dt_to and dt_from > dt_to:
        raise ValidationError("date_from must not be later than date_to.")
    return dt_from, dt_to


def date_filters(dt_from, dt_to, prefix):
    filters = {}
    if dt_from is not None:
        filters[f"{prefix}gte"] = dt_from
    if dt_to is not None:
        filters[f"{prefix}lte"] = dt_to
    return filters


def to_money(value):
    if value is None:
        value = Decimal("0.00")
    return f"{Decimal(value):.2f}"


def sales_metrics(business, dt_from, dt_to):
    sale_filter = date_filters(dt_from, dt_to, "created_at__")
    base = Sale.objects.filter(business=business, **sale_filter)

    total = base.count()
    draft = base.filter(status=Sale.Status.DRAFT).count()
    completed = base.filter(status=Sale.Status.COMPLETED).count()
    voided = base.filter(status=Sale.Status.VOIDED).count()

    revenue = SaleLine.objects.filter(
        sale__business=business,
        sale__status=Sale.Status.COMPLETED,
        **date_filters(dt_from, dt_to, "sale__created_at__"),
    ).aggregate(total=Sum(F("quantity") * F("unit_price")))["total"]
    revenue = Decimal(revenue) if revenue is not None else Decimal("0.00")

    loyalty = Sale.objects.filter(
        business=business,
        status=Sale.Status.COMPLETED,
        **sale_filter,
    ).aggregate(total=Sum("loyalty_earned"))["total"]
    loyalty = Decimal(loyalty) if loyalty is not None else Decimal("0.00")

    cogs = SaleLine.objects.filter(
        sale__business=business,
        sale__status=Sale.Status.COMPLETED,
        **date_filters(dt_from, dt_to, "sale__created_at__"),
    ).aggregate(
        total=Sum(
            F("quantity") * Coalesce(F("applied_cost_price"), Value(Decimal("0.00")))
        )
    )["total"]
    cogs = Decimal(cogs) if cogs is not None else Decimal("0.00")
    gross_profit = revenue - cogs

    return {
        "total": total,
        "completed": completed,
        "voided": voided,
        "draft": draft,
        "revenue": to_money(revenue),
        "cogs": to_money(cogs),
        "gross_profit": to_money(gross_profit),
        "loyalty_earned": to_money(loyalty),
    }


def purchasing_metrics(business, dt_from, dt_to):
    po_filter = date_filters(dt_from, dt_to, "created_at__")
    base = PurchaseOrder.objects.filter(business=business, **po_filter)
    total = base.count()
    draft = base.filter(status=PurchaseOrder.Status.DRAFT).count()
    confirmed = base.filter(status=PurchaseOrder.Status.CONFIRMED).count()
    cancelled = base.filter(status=PurchaseOrder.Status.CANCELLED).count()

    cost = PurchaseOrderLine.objects.filter(
        purchase_order__business=business,
        purchase_order__status=PurchaseOrder.Status.CONFIRMED,
        **date_filters(dt_from, dt_to, "purchase_order__created_at__"),
    ).aggregate(total=Sum(F("quantity") * F("unit_price")))["total"]
    cost = Decimal(cost) if cost is not None else Decimal("0.00")

    return {
        "total": total,
        "confirmed": confirmed,
        "cancelled": cancelled,
        "draft": draft,
        "cost": to_money(cost),
    }


def finance_metrics(business, dt_from, dt_to):
    expense_total = Expense.objects.filter(
        business=business,
        **date_filters(dt_from, dt_to, "created_at__"),
    ).aggregate(total=Sum("amount"))["total"]
    expense_total = Decimal(expense_total) if expense_total is not None else Decimal("0.00")

    sales = sales_metrics(business, dt_from, dt_to)
    revenue = Decimal(sales["revenue"]) if sales["revenue"] else Decimal("0.00")
    cogs = Decimal(sales["cogs"]) if sales["cogs"] else Decimal("0.00")
    gross_profit = revenue - cogs
    net_profit = gross_profit - expense_total

    journal_counts = (
        Journal.objects.filter(
            business=business,
            **date_filters(dt_from, dt_to, "created_at__"),
        )
        .values("status")
        .annotate(count=Count("id"))
    )
    journal_status = {"DRAFT": 0, "POSTED": 0, "REVERSED": 0}
    for row in journal_counts:
        journal_status[row["status"]] = row["count"]

    entry_filter = {
        "journal__business": business,
        "journal__status": Journal.Status.POSTED,
        **date_filters(dt_from, dt_to, "journal__created_at__"),
    }
    debit = JournalEntry.objects.filter(
        entry_type=JournalEntry.EntryType.DEBIT, **entry_filter
    ).aggregate(total=Sum("amount"))["total"]
    debit = Decimal(debit) if debit is not None else Decimal("0.00")

    credit = JournalEntry.objects.filter(
        entry_type=JournalEntry.EntryType.CREDIT, **entry_filter
    ).aggregate(total=Sum("amount"))["total"]
    credit = Decimal(credit) if credit is not None else Decimal("0.00")

    return {
        "expense_total": to_money(expense_total),
        "net_profit": to_money(net_profit),
        "journal": journal_status,
        "journal_entry": {
            "DEBIT": to_money(debit),
            "CREDIT": to_money(credit),
        },
    }


def counts_metrics(business):
    customers = Customer.objects.filter(business=business).count()
    products = Product.objects.filter(business=business).count()
    variants = Variant.objects.filter(product__business=business).count()
    employees = Employee.objects.filter(business=business).count()
    employees_active = Employee.objects.filter(business=business, active=True).count()
    return {
        "customers": customers,
        "products": products,
        "variants": variants,
        "employees": employees,
        "employees_active": employees_active,
    }


class OverviewView(BusinessAccessMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, business_id):
        business = self.require_business_permission("reports", "view")
        dt_from, dt_to = parse_date_params(request)
        return Response(
            {
                "sales": sales_metrics(business, dt_from, dt_to),
                "purchasing": purchasing_metrics(business, dt_from, dt_to),
                "finance": finance_metrics(business, dt_from, dt_to),
                "counts": counts_metrics(business),
            },
            status=200,
        )


class SalesReportView(BusinessAccessMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, business_id):
        business = self.require_business_permission("reports", "view")
        dt_from, dt_to = parse_date_params(request)
        return Response(sales_metrics(business, dt_from, dt_to), status=200)


class PurchasingReportView(BusinessAccessMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, business_id):
        business = self.require_business_permission("reports", "view")
        dt_from, dt_to = parse_date_params(request)
        return Response(purchasing_metrics(business, dt_from, dt_to), status=200)


class FinanceReportView(BusinessAccessMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, business_id):
        business = self.require_business_permission("reports", "view")
        dt_from, dt_to = parse_date_params(request)
        return Response(finance_metrics(business, dt_from, dt_to), status=200)


def inventory_metrics(business):
    from apps.inventory.models import Stock

    total_products = Product.objects.filter(business=business).count()
    total_variants = Variant.objects.filter(product__business=business).count()

    total_stock_quantity = Stock.objects.filter(location__business=business).aggregate(
        total=Sum("quantity")
    )["total"]
    if total_stock_quantity is None:
        total_stock_quantity = Decimal("0.00")

    variant_ids = Variant.objects.filter(product__business=business).values_list("id", flat=True)
    variant_qtys = (
        Stock.objects.filter(location__business=business)
        .values("variant_id")
        .annotate(total_qty=Sum("quantity"))
    )
    qty_map = {item["variant_id"]: item["total_qty"] or Decimal("0.00") for item in variant_qtys}

    low_stock_count = 0
    for v_id in variant_ids:
        qty = qty_map.get(v_id, Decimal("0.00"))
        if qty <= Decimal("5.00"):
            low_stock_count += 1

    inventory_value = Stock.objects.filter(location__business=business).aggregate(
        value=Sum(F("quantity") * F("variant__product__price"))
    )["value"]
    if inventory_value is None:
        inventory_value = Decimal("0.00")

    return {
        "total_products": total_products,
        "total_variants": total_variants,
        "total_stock_quantity": float(total_stock_quantity),
        "low_stock_count": low_stock_count,
        "inventory_value": f"{Decimal(inventory_value):.2f}",
    }


class InventoryReportView(BusinessAccessMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, business_id):
        business = self.require_business_permission("reports", "view")
        return Response(inventory_metrics(business), status=200)


def product_metrics(business):
    total_products = Product.objects.filter(business=business).count()
    active_products = total_products
    inactive_products = 0
    total_variants = Variant.objects.filter(product__business=business).count()
    variants_per_product = (
        Variant.objects.filter(product__business=business)
        .values("product__name")
        .annotate(variant_count=Count("id"))
    )
    summary = [
        {"name": row["product__name"], "variants": row["variant_count"]}
        for row in variants_per_product
    ]
    return {
        "total_products": total_products,
        "active_products": active_products,
        "inactive_products": inactive_products,
        "total_variants": total_variants,
        "summary": summary,
    }


def customer_metrics(business, dt_from, dt_to):
    total_customers = Customer.objects.filter(business=business).count()
    active_customers = total_customers
    sale_filter = date_filters(dt_from, dt_to, "created_at__")
    base = Sale.objects.filter(business=business, status=Sale.Status.COMPLETED, **sale_filter)
    by_customer = (
        base.values("customer__name")
        .annotate(count=Count("id"), revenue=Sum(F("lines__quantity") * F("lines__unit_price")))
        .order_by("-revenue")
    )
    top_customers = [
        {"name": row["customer__name"], "sales": row["count"], "revenue": to_money(row["revenue"])}
        for row in by_customer[:5]
        if row["customer__name"] is not None
    ]
    first_period = base.filter(**date_filters(dt_from, dt_to, "created_at__"))
    prev_from = None
    prev_to = None
    if dt_from and dt_to:
        prev_from = dt_from - (dt_to - dt_from)
        prev_to = dt_from - timedelta(microseconds=1)
    prev_period = Sale.objects.filter(business=business, status=Sale.Status.COMPLETED)
    if prev_from:
        prev_period = prev_period.filter(**date_filters(prev_from, prev_to, "created_at__"))
    customer_growth = {
        "current_period": total_customers,
        "previous_period": prev_period.count(),
        "growth": 0,
    }
    return {
        "total_customers": total_customers,
        "active_customers": active_customers,
        "customer_growth": customer_growth,
        "top_customers": top_customers,
    }


def supplier_metrics(business, dt_from, dt_to):
    total_suppliers = Supplier.objects.filter(business=business).count()
    active_suppliers = total_suppliers
    po_filter = date_filters(dt_from, dt_to, "created_at__")
    base = PurchaseOrder.objects.filter(business=business, **po_filter)
    purchase_volume = PurchaseOrderLine.objects.filter(
        purchase_order__business=business,
        **date_filters(dt_from, dt_to, "purchase_order__created_at__"),
    ).aggregate(total=Sum("quantity"))["total"]
    purchase_volume = Decimal(purchase_volume) if purchase_volume is not None else Decimal("0.00")
    purchase_value = PurchaseOrderLine.objects.filter(
        purchase_order__business=business,
        purchase_order__status=PurchaseOrder.Status.CONFIRMED,
        **date_filters(dt_from, dt_to, "purchase_order__created_at__"),
    ).aggregate(total=Sum(F("quantity") * F("unit_price")))["total"]
    purchase_value = Decimal(purchase_value) if purchase_value is not None else Decimal("0.00")
    per_supplier = (
        base.values("supplier__name")
        .annotate(po_count=Count("id"))
        .order_by("-po_count")
    )
    supplier_activity = [
        {"name": row["supplier__name"], "purchase_orders": row["po_count"]}
        for row in per_supplier
    ]
    return {
        "total_suppliers": total_suppliers,
        "active_suppliers": active_suppliers,
        "purchase_volume": float(purchase_volume),
        "purchase_value": to_money(purchase_value),
        "supplier_activity": supplier_activity,
    }


def promotion_metrics(business, dt_from, dt_to):
    total_promotions = Promotion.objects.filter(business=business).count()
    active_promotions = Promotion.objects.filter(business=business, status=Promotion.Status.ACTIVE).count()
    inactive_promotions = Promotion.objects.filter(business=business, status=Promotion.Status.INACTIVE).count()
    used_promotion_ids = SaleLine.objects.filter(
        sale__business=business,
        sale__status=Sale.Status.COMPLETED,
        applied_promotion__isnull=False,
        **date_filters(dt_from, dt_to, "sale__created_at__"),
    ).values_list("applied_promotion_id", flat=True)
    promotion_usage = (
        SaleLine.objects.filter(
            sale__business=business,
            sale__status=Sale.Status.COMPLETED,
            applied_promotion__isnull=False,
            **date_filters(dt_from, dt_to, "sale__created_at__"),
        )
        .values("applied_promotion__name")
        .annotate(redemption_count=Count("id"))
        .order_by("-redemption_count")
    )
    usage_list = [
        {"name": row["applied_promotion__name"], "redemption_count": row["redemption_count"]}
        for row in promotion_usage
    ]
    redemption_count = sum(row["redemption_count"] for row in promotion_usage)
    discount_summary = SaleLine.objects.filter(
        sale__business=business,
        sale__status=Sale.Status.COMPLETED,
        applied_promotion__isnull=False,
        **date_filters(dt_from, dt_to, "sale__created_at__"),
    ).aggregate(
        total_quantity=Sum("quantity"),
        total_discount=Sum("applied_discount_value"),
    )
    performance = [
        {
            "name": row["applied_promotion__name"],
            "discount_type": row["applied_discount_type"],
            "discount_value": to_money(row["discount_value"]),
        }
        for row in SaleLine.objects.filter(
            sale__business=business,
            sale__status=Sale.Status.COMPLETED,
            applied_promotion__isnull=False,
            **date_filters(dt_from, dt_to, "sale__created_at__"),
        )
        .values(
            "applied_promotion__name",
            "applied_discount_type",
            "applied_discount_value",
        )
        .distinct()
    ]
    return {
        "promotion_usage": usage_list,
        "redemption_count": redemption_count,
        "discount_summary": {
            "total_quantity": to_money(discount_summary["total_quantity"]),
            "total_discount": to_money(discount_summary["total_discount"]),
        },
        "performance": performance,
    }


def employee_metrics(business, dt_from, dt_to):
    total_count = Employee.objects.filter(business=business).count()
    active_count = Employee.objects.filter(business=business, active=True).count()
    sale_filter = date_filters(dt_from, dt_to, "created_at__")
    sales_agg = Sale.objects.filter(
        business=business, status=Sale.Status.COMPLETED, **sale_filter
    ).aggregate(
        sales_count=Count("id"),
    )
    revenue_agg = SaleLine.objects.filter(
        sale__business=business,
        sale__status=Sale.Status.COMPLETED,
        **date_filters(dt_from, dt_to, "sale__created_at__"),
    ).aggregate(total=Sum(F("quantity") * F("unit_price")))
    revenue = Decimal(revenue_agg["total"]) if revenue_agg["total"] is not None else Decimal("0.00")
    shift_agg = CashierShift.objects.filter(business=business).aggregate(
        total_shifts=Count("id"),
        completed_shifts=Count("status", filter=Q(status=CashierShift.Status.CLOSED)),
    )
    return {
        "employee_count": total_count,
        "active_employee_count": active_count,
        "employee_sales_summary": {
            "sales_count": sales_agg["sales_count"] or 0,
            "revenue": to_money(revenue),
        },
        "shift_activity": {
            "total_shifts": shift_agg["total_shifts"] or 0,
            "completed_shifts": shift_agg["completed_shifts"] or 0,
        },
    }


def node19_report_payload(report_type, business, dt_from, dt_to):
    if report_type == "product":
        return product_metrics(business)
    return None


class ProductReportView(BusinessAccessMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, business_id):
        business = self.require_business_permission("reports", "view")
        dt_from, dt_to = parse_date_params(request)
        return Response(product_metrics(business), status=200)


class CustomerReportView(BusinessAccessMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, business_id):
        business = self.require_business_permission("reports", "view")
        dt_from, dt_to = parse_date_params(request)
        return Response(customer_metrics(business, dt_from, dt_to), status=200)


class SupplierReportView(BusinessAccessMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, business_id):
        business = self.require_business_permission("reports", "view")
        dt_from, dt_to = parse_date_params(request)
        return Response(supplier_metrics(business, dt_from, dt_to), status=200)


class PromotionReportView(BusinessAccessMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, business_id):
        business = self.require_business_permission("reports", "view")
        dt_from, dt_to = parse_date_params(request)
        return Response(promotion_metrics(business, dt_from, dt_to), status=200)


class EmployeeReportView(BusinessAccessMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, business_id):
        business = self.require_business_permission("reports", "view")
        dt_from, dt_to = parse_date_params(request)
        return Response(employee_metrics(business, dt_from, dt_to), status=200)


def _build_csv_response(data, filename):
    response = HttpResponse(content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    writer = csv.writer(response)
    _flatten_to_csv(writer, data)
    return response


def _build_xlsx_response(data, filename):
    if not OPENPYXL_AVAILABLE:
        raise ValidationError("XLSX export not available. Please install openpyxl.")
    wb = Workbook()
    ws = wb.active
    _flatten_to_xlsx(ws, data)
    bio = BytesIO()
    wb.save(bio)
    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _flatten_to_csv(writer, data, parent_key=""):
    if isinstance(data, dict):
        for key, value in data.items():
            new_key = f"{parent_key}.{key}" if parent_key else key
            if isinstance(value, (dict, list)):
                _flatten_to_csv(writer, value, new_key)
            else:
                writer.writerow([new_key, value])
    elif isinstance(data, list):
        for idx, item in enumerate(data):
            new_key = f"{parent_key}[{idx}]"
            if isinstance(item, (dict, list)):
                _flatten_to_csv(writer, item, new_key)
            else:
                writer.writerow([new_key, item])
    else:
        writer.writerow([parent_key, data])


def _flatten_to_xlsx(ws, data, parent_key="", row=1):
    if isinstance(data, dict):
        for key, value in data.items():
            new_key = f"{parent_key}.{key}" if parent_key else key
            if isinstance(value, (dict, list)):
                row = _flatten_to_xlsx(ws, value, new_key, row)
            else:
                ws.cell(row=row, column=1, value=new_key)
                ws.cell(row=row, column=2, value=value)
                row += 1
    elif isinstance(data, list):
        for idx, item in enumerate(data):
            new_key = f"{parent_key}[{idx}]"
            if isinstance(item, (dict, list)):
                row = _flatten_to_xlsx(ws, item, new_key, row)
            else:
                ws.cell(row=row, column=1, value=new_key)
                ws.cell(row=row, column=2, value=item)
                row += 1
    else:
        ws.cell(row=row, column=1, value=parent_key)
        ws.cell(row=row, column=2, value=data)
        row += 1
    return row


class ExportReportView(BusinessAccessMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, business_id, report_type, fmt):
        business = self.require_business_permission("reports", "view")
        dt_from, dt_to = parse_date_params(request)

        data = None
        if report_type == "overview":
            data = {
                "sales": sales_metrics(business, dt_from, dt_to),
                "purchasing": purchasing_metrics(business, dt_from, dt_to),
                "finance": finance_metrics(business, dt_from, dt_to),
                "counts": counts_metrics(business),
            }
        elif report_type == "sales":
            data = sales_metrics(business, dt_from, dt_to)
        elif report_type == "purchasing":
            data = purchasing_metrics(business, dt_from, dt_to)
        elif report_type == "finance":
            data = finance_metrics(business, dt_from, dt_to)
        elif report_type == "inventory":
            data = inventory_metrics(business)
        elif report_type == "product":
            data = product_metrics(business)
        elif report_type == "customer":
            data = customer_metrics(business, dt_from, dt_to)
        elif report_type == "supplier":
            data = supplier_metrics(business, dt_from, dt_to)
        elif report_type == "promotion":
            data = promotion_metrics(business, dt_from, dt_to)
        elif report_type == "employee":
            data = employee_metrics(business, dt_from, dt_to)

        if data is None:
            raise ValidationError(f"Unknown report type: {report_type}")

        filename = f"{report_type}_{business_id}.{fmt}"

        if fmt == "csv":
            return _build_csv_response(data, filename)
        elif fmt == "xlsx":
            return _build_xlsx_response(data, filename)
        else:
            raise ValidationError("Unsupported format. Use csv or xlsx.")


# =============================================================================
# GAP-04DASH-CASHFLOW — Operational Cash Flow Report
# =============================================================================

class CashflowView(BusinessAccessMixin, APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, business_id):
        business = self.require_business_permission("reports", "view")
        dt_from, dt_to = parse_date_params(request)
        location_id = request.query_params.get("location")

        from apps.receivable.models import Receivable, PaymentAllocation
        from apps.payable.models import Payable, SupplierPaymentAllocation

        sale_filter = date_filters(dt_from, dt_to, "created_at__")
        location_filter = {}
        if location_id:
            loc = get_object_or_404(Location, pk=location_id, business=business)
            location_filter = {"location": loc}

        # =========================================================================
        # INFLOW: A. Regular non-credit POS sales
        # =========================================================================
        pos_sales_filter = {
            "business": business,
            "status": Sale.Status.COMPLETED,
            "receivable__isnull": True,
            "payment_method__in": ["CASH", "QRIS", "TRANSFER"],
            **sale_filter,
        }
        if location_filter:
            pos_sales_filter["location"] = location_filter["location"]

        pos_cash_sales = SaleLine.objects.filter(
            sale__in=Sale.objects.filter(**pos_sales_filter),
            **date_filters(dt_from, dt_to, "sale__created_at__"),
        ).aggregate(
            total=Sum(F("quantity") * F("unit_price"))
        )["total"]
        pos_cash_sales = Decimal(pos_cash_sales) if pos_cash_sales is not None else Decimal("0.00")

        # =========================================================================
        # INFLOW: B. Receivable collections (GAP-02 PaymentAllocation)
        # Includes BOTH valid (is_reversed=False) AND reversed allocations
        # because reversed allocation creates its own reversal movement
        # =========================================================================
        receivable_alloc_filter = {"receivable__business": business}
        if location_filter:
            receivable_alloc_filter["receivable__location"] = location_filter["location"]
        if dt_from:
            receivable_alloc_filter["payment_date__gte"] = dt_from
        if dt_to:
            receivable_alloc_filter["payment_date__lte"] = dt_to

        # Also include reversed allocations for the reversal movement
        receivable_alloc_all = PaymentAllocation.objects.filter(**receivable_alloc_filter)

        # =========================================================================
        # OUTFLOW: A. Supplier payments (GAP-03 SupplierPaymentAllocation)
        # =========================================================================
        supplier_alloc_filter = {"payable__business": business}
        if location_filter:
            supplier_alloc_filter["payable__location"] = location_filter["location"]
        if dt_from:
            supplier_alloc_filter["payment_date__gte"] = dt_from
        if dt_to:
            supplier_alloc_filter["payment_date__lte"] = dt_to

        supplier_alloc_all = SupplierPaymentAllocation.objects.filter(**supplier_alloc_filter)

        # =========================================================================
        # OUTFLOW: B. Expenses
        # =========================================================================
        expense_filter = {"business": business}
        if location_filter:
            # Expenses don't have location FK, so skip location filter
            pass
        if dt_from:
            expense_filter["created_at__gte"] = dt_from
        if dt_to:
            expense_filter["created_at__lte"] = dt_to

        expense_total = Expense.objects.filter(**expense_filter).aggregate(
            total=Sum("amount")
        )["total"]
        expense_total = Decimal(expense_total) if expense_total is not None else Decimal("0.00")

        # =========================================================================
        # Build cash_movements list
        # =========================================================================
        movements = []

        # A. POS Cash Sales inflows
        for sale in Sale.objects.filter(**pos_sales_filter):
            line_total = SaleLine.objects.filter(sale=sale).aggregate(
                t=Sum(F("quantity") * F("unit_price"))
            )["t"]
            line_total = Decimal(line_total) if line_total is not None else Decimal("0.00")
            if line_total > 0:
                movements.append({
                    "id": str(sale.id),
                    "date": sale.created_at.isoformat(),
                    "direction": "INFLOW",
                    "source_type": "POS_SALE",
                    "reference": f"Sale {str(sale.id)[:8]}",
                    "payment_method": sale.payment_method,
                    "amount": to_money(line_total),
                    "is_reversal": False,
                })

        # B. Receivable PaymentAllocation movements (valid + reversal)
        for alloc in receivable_alloc_all:
            if alloc.payment_method not in ("CASH", "QRIS", "TRANSFER"):
                continue
            if not alloc.is_reversed:
                # Normal inflow
                movements.append({
                    "id": str(alloc.id),
                    "date": alloc.payment_date.isoformat() if alloc.payment_date else "",
                    "direction": "INFLOW",
                    "source_type": "RECEIVABLE_PAYMENT",
                    "reference": f"Allocation {str(alloc.id)[:8]}",
                    "payment_method": alloc.payment_method,
                    "amount": to_money(alloc.amount),
                    "is_reversal": False,
                })
            else:
                # Original movement remains historical
                if alloc.payment_date:
                    movements.append({
                        "id": str(alloc.id),
                        "date": alloc.payment_date.isoformat(),
                        "direction": "INFLOW",
                        "source_type": "RECEIVABLE_PAYMENT",
                        "reference": f"Allocation {str(alloc.id)[:8]}",
                        "payment_method": alloc.payment_method,
                        "amount": to_money(alloc.amount),
                        "is_reversal": False,
                    })
                # Reversal movement
                if alloc.reversed_at:
                    movements.append({
                        "id": f"{alloc.id}-reversal",
                        "date": alloc.reversed_at.isoformat(),
                        "direction": "INFLOW_REVERSAL",
                        "source_type": "RECEIVABLE_PAYMENT",
                        "reference": f"Reversal of Allocation {str(alloc.id)[:8]}",
                        "payment_method": alloc.payment_method,
                        "amount": to_money(Decimal("0.00") - alloc.amount),
                        "is_reversal": True,
                    })
                elif alloc.is_reversed:
                    # reversed_at is null but is_reversed=True
                    # Use updated_at as fallback timestamp
                    if alloc.updated_at:
                        movements.append({
                            "id": f"{alloc.id}-reversal",
                            "date": alloc.updated_at.isoformat(),
                            "direction": "INFLOW_REVERSAL",
                            "source_type": "RECEIVABLE_PAYMENT",
                            "reference": f"Reversal of Allocation {str(alloc.id)[:8]}",
                            "payment_method": alloc.payment_method,
                            "amount": to_money(Decimal("0.00") - alloc.amount),
                            "is_reversal": True,
                        })

        # C. Supplier PaymentAllocation movements (outflow)
        for alloc in supplier_alloc_all:
            if alloc.payment_method not in ("CASH", "QRIS", "TRANSFER"):
                continue
            if not alloc.is_reversed:
                movements.append({
                    "id": str(alloc.id),
                    "date": alloc.payment_date.isoformat() if alloc.payment_date else "",
                    "direction": "OUTFLOW",
                    "source_type": "SUPPLIER_PAYMENT",
                    "reference": f"Supplier Allocation {str(alloc.id)[:8]}",
                    "payment_method": alloc.payment_method,
                    "amount": to_money(alloc.amount),
                    "is_reversal": False,
                })
            else:
                if alloc.payment_date:
                    movements.append({
                        "id": str(alloc.id),
                        "date": alloc.payment_date.isoformat(),
                        "direction": "OUTFLOW",
                        "source_type": "SUPPLIER_PAYMENT",
                        "reference": f"Supplier Allocation {str(alloc.id)[:8]}",
                        "payment_method": alloc.payment_method,
                        "amount": to_money(alloc.amount),
                        "is_reversal": False,
                    })
                if alloc.reversed_at:
                    movements.append({
                        "id": f"{alloc.id}-reversal",
                        "date": alloc.reversed_at.isoformat(),
                        "direction": "OUTFLOW_REVERSAL",
                        "source_type": "SUPPLIER_PAYMENT",
                        "reference": f"Reversal of Supplier Allocation {str(alloc.id)[:8]}",
                        "payment_method": alloc.payment_method,
                        "amount": to_money(Decimal("0.00") - alloc.amount),
                        "is_reversal": True,
                    })
                elif alloc.is_reversed and alloc.updated_at:
                    movements.append({
                        "id": f"{alloc.id}-reversal",
                        "date": alloc.updated_at.isoformat(),
                        "direction": "OUTFLOW_REVERSAL",
                        "source_type": "SUPPLIER_PAYMENT",
                        "reference": f"Reversal of Supplier Allocation {str(alloc.id)[:8]}",
                        "payment_method": alloc.payment_method,
                        "amount": to_money(Decimal("0.00") - alloc.amount),
                        "is_reversal": True,
                    })

        # D. Expense movements
        for expense in Expense.objects.filter(**expense_filter):
            movements.append({
                "id": str(expense.id),
                "date": expense.created_at.isoformat(),
                "direction": "OUTFLOW",
                "source_type": "EXPENSE",
                "reference": expense.description[:50] if expense.description else "",
                "payment_method": "",
                "amount": to_money(expense.amount),
                "is_reversal": False,
            })

        # Sort by date descending
        movements.sort(key=lambda m: m.get("date", ""), reverse=True)

        # =========================================================================
        # Calculate totals
        # =========================================================================
        # Receiverable collection totals
        collections_valid = receivable_alloc_all.filter(is_reversed=False).aggregate(
            total=Sum("amount")
        )["total"]
        collections_valid = Decimal(collections_valid) if collections_valid is not None else Decimal("0.00")

        # Supplier payment totals
        supplier_valid = supplier_alloc_all.filter(is_reversed=False).aggregate(
            total=Sum("amount")
        )["total"]
        supplier_valid = Decimal(supplier_valid) if supplier_valid is not None else Decimal("0.00")

        total_inflow = pos_cash_sales + collections_valid
        total_outflow = supplier_valid + expense_total
        net_cashflow = total_inflow - total_outflow

        return Response({
            "summary": {
                "total_inflow": to_money(total_inflow),
                "total_outflow": to_money(total_outflow),
                "net_cashflow": to_money(net_cashflow),
            },
            "inflow_breakdown": {
                "pos_cash_sales": to_money(pos_cash_sales),
                "receivable_collections": to_money(collections_valid),
            },
            "outflow_breakdown": {
                "supplier_payments": to_money(supplier_valid),
                "expenses": to_money(expense_total),
            },
            "cash_movements": movements,
        }, status=drf_status.HTTP_200_OK)
